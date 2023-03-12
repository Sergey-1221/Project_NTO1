import openpyxl


def get_excel(name_xl, num, i_min, col):
    # Открываю книгу
    wb = openpyxl.load_workbook(name_xl)
    # Открываю страницу
    sheet = wb.get_sheet_by_name(f'{wb.get_sheet_names()[num-1]}')
    # Иду по циклу, пока не встречу пустую строку
    while not(sheet.cell(row=i_min, column=1).value is None):
        for j in range(1, col):
            print(sheet.cell(row=i_min, column=j).value, end=' ')
        i_min += 1
        print()


if __name__ == '__main__':
    # Первый аргумент - страница, второй - с какой строки читать, третий - до какого столбца
    get_excel("Tours.xlsx", 1, 3, 6)


